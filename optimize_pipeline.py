import numpy as np
from colormath.color_objects import LabColor, XYZColor, sRGBColor
from colormath.color_conversions import convert_color
from Spectrum_Functions import spec_to_xyz
from Spectrum_Functions import delta_E_CIE2000
from Spectrum_Functions import gen_spectrum_1dip
from Spectrum_Functions import gen_spectrum_2dip
from Spectrum_Functions import gen_spectrum_1gauss
from Spectrum_Functions import gen_spectrum_2gauss
from scipy.optimize import minimize
from scipy.optimize import brute
import xlsxwriter
import openpyxl
import os
import matplotlib.pyplot as plt
from sys import getsizeof
from solcore.optimization import DE, PDE
from solcore.light_source import LightSource
from solcore.solar_cell import SolarCell
from solcore.solar_cell_solver import solar_cell_solver
from solcore.structure import Junction

from colormath.color_objects import LabColor, XYZColor,  sRGBColor
from colormath.color_conversions import convert_color
import pandas as pd
import random
from skimage import io

interval=0.5 # interval between each two wavelength points, 0.02 needed for low dE values
wl=np.arange(380,780+interval,interval)

wl_cell=np.arange(300, 4000, 0.5)

df = pd.read_excel("ASTMG173_split.xlsx", sheet_name=0)

light = LightSource(
    source_type="standard", version="AM1.5g", x=wl_cell, output_units="photon_flux_per_nm"
)

class Optimize:
    def __init__(self, colors_tuple, color_name):
        '''
        Send color in from op_send_sq.py script
        colors_tuple is (a,b,c) LAB coordinates
        color_name is the name of the color from 2005 ColorChecker, used to save the filename
        
        Function calculates uses scipy minimize to find the optimum deltaE value.
        Reads in values from excel sheets generated from matlab
        Saves data out to excel sheet

        '''
        print(color_name)
        colors_tuple=tuple(colors_tuple)
        self.lab = LabColor(colors_tuple[0], colors_tuple[1], colors_tuple[2])
        self.xyz = convert_color(self.lab, XYZColor)
        self.rgb = convert_color(self.xyz, sRGBColor)
        self.red = self.rgb.get_value_tuple()[0]
        self.green = self.rgb.get_value_tuple()[1]
        self.blue = self.rgb.get_value_tuple()[2]


        self.peak_ = 1
        self.base_ = 0

        # for writing out to excel and calculating delta E
        self.color = str(self.red)+'_'+str(self.green)+'_'+str(self.blue)
        print('red ',self.red,' green ',self.green,' blue ',self.blue )
        self.target_color_sRGB=sRGBColor(self.red,self.green,self.blue) # RGB range is 0 to 1
        self.target_color_LAB=convert_color(self.target_color_sRGB, LabColor)
        self.target_color_LAB=(self.target_color_LAB.lab_l,self.target_color_LAB.lab_a,self.target_color_LAB.lab_b)
        print('The Target Color is ', self.target_color_LAB)
        self.color_lab = str(self.target_color_LAB[0])+'_'+str(self.target_color_LAB[1])+'_'+str(self.target_color_LAB[2]) 
        
        self.color_2005 = color_name #name from ColorChecker 2005
        # define ranges for optimize
        self.vis_bound = [380,780]
        self.width_max = [0.0,300]
        self.width2_max = [0.0,150]
        self.peak_range = [1,1]
        
        self.ranges=[self.vis_bound,self.width_max,self.peak_range]
        self.ranges2d = [self.vis_bound,self.width2_max,self.peak_range,self.vis_bound,self.width2_max, self.peak_range]

        self.ranges_MJ= [self.vis_bound,self.width_max,self.peak_range, [0.8, 1.2], [1.7, 2.1]]
        self.ranges2d_MJ = [self.vis_bound, self.width2_max, self.peak_range, self.vis_bound, self.width2_max, self.peak_range, [0.5, 1.4], [1.4, 2]]

        # self.x0 = [random.uniform(380,780),random.uniform(0,400), random.uniform(0,1)]
        # self.x02d = [random.uniform(380,780),random.uniform(0,200),random.uniform(380,780),random.uniform(0,200)]

        # self.x0 = [520, 100, 0.5]
        # self.x02d = [400, 100, 700, 100]

        self.d1_cp  = {} #  spectrum, width, center, peak color parameters to be used for each delta E to calculate color spectrum
        self.d2_cp  = {}
        self.g1_cp  = {}
        self.g2_cp  = {}
        self.dE_threshold1d = 2
        self.dE_threshold = 2
        
        method1 = 'trust-constr'
        method2 = 'Nelder-Mead'

        # print('Initial', self.x0, self.x02d)
        # tolerance = 1e-9 #100 1e-9 #0.001
        
        # self.result1d_Minimize = brute(self.op_delta_E_1dip, ranges=self.ranges)
        # self.result1gauss_Minimize = brute(self.op_delta_E_1gauss, ranges=self.ranges)
        # self.result2d_Minimize = brute(self.op_delta_E_2dip, ranges=self.ranges2d)
        # self.result2gauss_Minimize = brute(self.op_delta_E_2gauss, ranges=self.ranges2d)


        print("Optimization for 1 top hat dip")
        # self.result1d_Minimize = minimize(self.op_delta_E_1dip,self.x0,method= method2,
        #                                   bounds=self.ranges,tol = tolerance,
        #                                   args=df)#,options={'return_all': True,'fatol': 0.1,'adaptive': True})#'initial_tr_radius': 0.5})

        PDE_obj_1dip = PDE(self.op_delta_E_1dip, mutation=(0.5, 2),
                                     bounds=self.ranges, maxiters=100)

        self.result1d_Minimize = PDE_obj_1dip.solve()
        res = self.result1d_Minimize
        best_pop_1d = res[0]

        print('parameters for best result:', best_pop_1d, '\n', 'optimized value:', res[1])

        print("Optimization for 1 top hat dip (2)")
        # self.result1d_Minimize = minimize(self.op_delta_E_1dip,self.x0,method= method2,
        #                                   bounds=self.ranges,tol = tolerance,
        #                                   args=df)#,options={'return_all': True,'fatol': 0.1,'adaptive': True})#'initial_tr_radius': 0.5})

        elec_1d = op_1dip_MJ(self.target_color_LAB, self.dE_threshold1d)

        PDE_obj_1dip = PDE(elec_1d.evaluate, mutation=(0.5, 2),
                                     bounds=self.ranges_MJ, maxiters=75)


        self.result1d_Minimize = PDE_obj_1dip.solve()
        res = self.result1d_Minimize
        best_pop_1d = res[0]

        elec_1d.plot(best_pop_1d)

        print('parameters for best result:', best_pop_1d, '\n', 'optimized value:', res[1])


        # best_pop_evo = res[2]
        # best_fitn_evo = res[3]
        # mean_fitn_evo = res[4]
        # final_fitness = res[1]

        # plot evolution of the fitness of the best population per iteration

        # plt.figure()
        # plt.plot(best_fitn_evo, '-k', label='Best')
        # plt.plot(mean_fitn_evo, '--r', label='Mean')
        # plt.xlabel('iteration')
        # plt.ylabel('delta_E')
        # plt.title('1 dip')
        # plt.ylim(-1,0)
        # plt.show()

        print("Optimization for 1 Gaussian dip")
        # self.result1gauss_Minimize = minimize(self.op_delta_E_1gauss,self.x0,method=method2,
        #                                       bounds=self.ranges,tol = tolerance,
        #                                       args=df)

        PDE_obj_1gauss = DE(self.op_delta_E_1gauss, mutation=(0.5, 2),
                          bounds=self.ranges, maxiters=75)

        self.result1gauss_Minimize = PDE_obj_1gauss.solve()
        res = self.result1gauss_Minimize
        best_pop_1g = res[0]

        print('parameters for best result:', best_pop_1g, '\n', 'optimized value:', res[1])

        best_pop_evo = res[2]
        best_fitn_evo = res[3]
        mean_fitn_evo = res[4]
        final_fitness = res[1]

        # plot evolution of the fitness of the best population per iteration
        #
        # plt.figure()
        # plt.plot(best_fitn_evo, '-k', label='Best')
        # plt.plot(mean_fitn_evo, '--r', label='Mean')
        # plt.xlabel('iteration')
        # plt.ylabel('delta_E')
        # plt.title('1 gaussian')
        # plt.show()

        print("Optimization for 2 top hat dips")
        # self.result2d_Minimize = minimize(self.op_delta_E_2dip,self.x02d,method=method2,
        #                                   bounds=self.ranges2d,tol = tolerance,
        #                                   args=df)

        PDE_obj_2dip = DE(self.op_delta_E_2dip, mutation=(0.5, 2),
                                     bounds=self.ranges2d_MJ, maxiters=1000)

        self.result2d_Minimize = PDE_obj_2dip.solve()
        res = self.result2d_Minimize
        best_pop_2d = res[0]

        print('parameters for best result:', best_pop_2d, '\n', 'optimized value:', res[1])

        best_pop_evo = res[2]
        best_fitn_evo = res[3]
        mean_fitn_evo = res[4]
        final_fitness = res[1]

        # plot evolution of the fitness of the best population per iteration

        # plt.figure()
        # plt.plot(best_fitn_evo, '-k', label='Best')
        # plt.plot(mean_fitn_evo, '--r', label='Mean')
        # plt.xlabel('iteration')
        # plt.ylabel('delta_E')
        # plt.title('2 dip')
        # plt.show()

        print("Optimization for 2 top hat dip (2)")
        # self.result1d_Minimize = minimize(self.op_delta_E_1dip,self.x0,method= method2,
        #                                   bounds=self.ranges,tol = tolerance,
        #                                   args=df)#,options={'return_all': True,'fatol': 0.1,'adaptive': True})#'initial_tr_radius': 0.5})

        elec_2d = op_2dip_MJ(self.target_color_LAB, self.dE_threshold)

        PDE_obj_2dip = PDE(elec_2d.evaluate, mutation=(0.5, 2),
                                     bounds=self.ranges2d_MJ, maxiters=150)


        self.result2d_Minimize = PDE_obj_2dip.solve()
        res = self.result2d_Minimize
        best_pop_2d = res[0]

        elec_2d.plot(best_pop_2d)

        print('parameters for best result:', best_pop_1d, '\n', 'optimized value:', res[1])


        print("Optimization for 2 Gaussian dips")
        # self.result2gauss_Minimize = minimize(self.op_delta_E_2gauss,self.x02d,method=method2,
        #                                       bounds=self.ranges2d,tol = tolerance,
        #                                       args=df)

        PDE_obj_2gauss = DE(self.op_delta_E_2gauss, mutation=(0.5, 2),
                            bounds=self.ranges2d, maxiters=300)

        self.result2gauss_Minimize = PDE_obj_2gauss.solve()
        res = self.result2gauss_Minimize
        best_pop_2g = res[0]

        print('parameters for best result:', best_pop_2g, '\n', 'optimized value:', res[1])

        best_pop_evo = res[2]
        best_fitn_evo = res[3]
        mean_fitn_evo = res[4]
        final_fitness = res[1]

        # plot evolution of the fitness of the best population per iteration

        # plt.figure()
        # plt.plot(best_fitn_evo, '-k', label='Best')
        # plt.plot(mean_fitn_evo, '--r', label='Mean')
        # plt.xlabel('iteration')
        # plt.ylabel('delta_E')
        # plt.title('2 gaussian')
        # plt.show()

        best_pop1d = best_pop_1d[:3]
        best_pop2d = best_pop_2d[:6]
        plt.figure()
        plt.plot(wl, gen_spectrum_1dip(*best_pop1d))
        plt.plot(wl, gen_spectrum_1gauss(*best_pop_1g))
        plt.plot(wl, gen_spectrum_2dip(*best_pop2d))
        plt.plot(wl, gen_spectrum_2gauss(*best_pop_2g))
        plt.title(color_name)
        plt.show()

        rgb_target = convert_color(self.xyz, sRGBColor)
        rgb_target_list = [rgb_target.rgb_r*255, rgb_target.rgb_g*255, rgb_target.rgb_b*255]

        XYZ=spec_to_xyz(gen_spectrum_1dip(*best_pop1d), df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        rgb_d1 = convert_color(XYZ, sRGBColor)
        rgb_d1_list = [rgb_d1.rgb_r*255, rgb_d1.rgb_g*255, rgb_d1.rgb_b*255]

        XYZ=spec_to_xyz(gen_spectrum_1gauss(*best_pop_1g), df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        rgb_g1 = convert_color(XYZ, sRGBColor)
        rgb_g1_list = [rgb_g1.rgb_r*255, rgb_g1.rgb_g*255, rgb_g1.rgb_b*255]

        XYZ=spec_to_xyz(gen_spectrum_2dip(*best_pop2d), df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        rgb_d2 = convert_color(XYZ, sRGBColor)
        rgb_d2_list = [rgb_d2.rgb_r*255, rgb_d2.rgb_g*255, rgb_d2.rgb_b*255]

        XYZ=spec_to_xyz(gen_spectrum_2gauss(*best_pop_2g), df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        rgb_g2 = convert_color(XYZ, sRGBColor)
        rgb_g2_list = [rgb_g2.rgb_r*255, rgb_g2.rgb_g*255, rgb_g2.rgb_b*255]

        print(rgb_target_list)
        palette = np.array([[rgb_target_list, rgb_target_list],  # index 1: green
                            [rgb_d1_list, rgb_g1_list],  # index 3: white
                            [rgb_d2_list, rgb_g2_list]], dtype=int)  # index 5: yellow)


        io.imshow(palette)
        plt.show()

        self.tu_cp = (self.d1_cp, self.d2_cp, self.g1_cp, self.g2_cp) #tuple of color parameter dictionaries to easily access all of these values at once
        #DEFINE WCDE DICT
        
        return


    def op_delta_E_1dip(self,x, df=df):
    
        center = x[0]
        width = x[1]
        peak = x[2]
        #base = x[3]
        spectrum = gen_spectrum_1dip(center,width,peak,base=0)
        XYZ=spec_to_xyz(spectrum, df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        Lab=convert_color(XYZ, LabColor)
        delta_E=delta_E_CIE2000((Lab.lab_l,Lab.lab_a,Lab.lab_b),self.target_color_LAB)
        if delta_E < self.dE_threshold1d:
            mean_R = np.mean(spectrum)
            scaled_E = mean_R*delta_E # penalize high average R
        else:
            scaled_E = delta_E
        #print('delta_E 1dip is ',delta_E)
        # if delta_E < self.dE_threshold1d:
        #     for i in range(2): # Trying to avoid excel writing error
        #         #print('Delta E Below Threshold')
        #         #WRITE EXCEL FUNCTION
        #         xcel_file = self.color_2005+'_D1_'+'dE_'+str(delta_E).replace('.','_')#color_name
        #         cwd = os.getcwd()
        #         filepath=os.path.join(cwd, r"color_data",xcel_file+'.xlsx')
        #         cp_list =([center, width, peak, spectrum,Lab,filepath]) # color parameter list to be stored in delta_E dict for op send to try different color parameters
        #         self.d1_cp[delta_E] = cp_list
        #
        #         workbook = xlsxwriter.Workbook(filepath)
        #         worksheet = workbook.add_worksheet()
        #         worksheet.write('A1', 'Efficiency')
        #         worksheet.write('B1', 'dE_D1')
        #         worksheet.write('B2', delta_E)
        #         worksheet.write('C1', 'RGB color')
        #         worksheet.write('C2',self.color)
        #         worksheet.write('D1','LAB color')
        #         worksheet.write('D2', self.color_lab)
        #         worksheet.write('E1', 'Eg')
        #         worksheet.write('F1', 'eff_CM')
        #
        #         worksheet.write('G1', 'Color')
        #         worksheet.write('G2', self.color_2005)
        #         worksheet.write('H1', 'Method')
        #         worksheet.write('H2', 'D1')
        #
        #
        #         worksheet.write('I1', 'Center')
        #         worksheet.write('I2', center)
        #         worksheet.write('J1', 'Width')
        #         worksheet.write('J2', width)
        #         worksheet.write('K1', 'Peak')
        #         worksheet.write('K2', peak)
        #
        #         workbook.close()
        # return delta_E
        return scaled_E
    
    def op_delta_E_1gauss(self,x, df=df):


        center = x[0]
        width = x[1]
        peak = x[2]
        spectrum = gen_spectrum_1gauss(center,width,peak,base=0)
        XYZ=spec_to_xyz(spectrum, df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        Lab=convert_color(XYZ, LabColor)

        delta_E=delta_E_CIE2000((Lab.lab_l,Lab.lab_a,Lab.lab_b),self.target_color_LAB)

        if delta_E < self.dE_threshold1d:
            mean_R = np.mean(spectrum)
            scaled_E = mean_R*delta_E # penalize high average R
        else:
            scaled_E = delta_E
        #print('delta_E 1gauss is ',delta_E)
        # if delta_E < self.dE_threshold1d:
        #     for i in range(2):
        #         #print('Delta E Below Threshold')
        #         xcel_file = self.color_2005+'_G1_'+'dE_'+str(delta_E).replace('.','_')#color_name
        #         cwd = os.getcwd()
        #         filepath=os.path.join(cwd, r"color_data",xcel_file+'.xlsx')
        #         cp_list =([center, width, peak, spectrum,Lab,filepath])
        #         self.g1_cp[delta_E] = cp_list
        #
        #         workbook = xlsxwriter.Workbook(filepath)
        #         worksheet = workbook.add_worksheet()
        #         worksheet.write('A1', 'Efficiency')
        #         worksheet.write('B1', 'dE_G1')
        #         worksheet.write('B2', delta_E)
        #         worksheet.write('C1', 'RGB color')
        #         worksheet.write('C2',self.color)
        #         worksheet.write('D1','LAB color')
        #         worksheet.write('D2', self.color_lab)
        #         worksheet.write('E1', 'Eg')
        #         worksheet.write('F1', 'eff_CM')
        #
        #         worksheet.write('G1', 'Color')
        #         worksheet.write('G2', self.color_2005)
        #         worksheet.write('H1', 'Method')
        #         worksheet.write('H2', 'G1')
        #
        #
        #         worksheet.write('I1', 'Center')
        #         worksheet.write('I2', center)
        #         worksheet.write('J1', 'Width')
        #         worksheet.write('J2', width)
        #         worksheet.write('K1', 'Peak')
        #         worksheet.write('K2', peak)
        #         workbook.close()
        # return delta_E
        return scaled_E
    
    
    def op_delta_E_2dip(self,x, df=df):
        center = x[0]
        width = x[1]
        peak1 = x[2]
        center2 = x[3]
        width2 = x[4]
        peak2 = x[5]
        spectrum = gen_spectrum_2dip(center,width,peak1,center2,width2,peak2)
        XYZ=spec_to_xyz(spectrum, df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        Lab=convert_color(XYZ, LabColor)

        delta_E=delta_E_CIE2000((Lab.lab_l,Lab.lab_a,Lab.lab_b),self.target_color_LAB)
        if delta_E < self.dE_threshold:
            mean_R = np.mean(spectrum)
            scaled_E = mean_R*delta_E # penalize high average R
        else:
            scaled_E = delta_E
        #print('delta_E 2dip is ',delta_E)
        # if delta_E < self.dE_threshold:
        #     for i in range(2):
        #         xcel_file = self.color_2005+'_D2_'+'dE_'+str(delta_E).replace('.','_')#color_name
        #         cwd = os.getcwd()
        #         filepath=os.path.join(cwd, r"color_data",xcel_file+'.xlsx')
        #
        #         #print('Delta E Below Threshold')
        #         cp_list = ([center, width, center2, width2,spectrum,Lab,filepath])
        #         self.d2_cp[delta_E] = cp_list
        #
        #
        #         workbook = xlsxwriter.Workbook(filepath)
        #         worksheet = workbook.add_worksheet()
        #         worksheet.write('A1', 'Efficiency')
        #         worksheet.write('B1', 'dE_D2')
        #         worksheet.write('B2', delta_E)
        #         worksheet.write('C1', 'RGB color')
        #         worksheet.write('C2',self.color)
        #         worksheet.write('D1','LAB color')
        #         worksheet.write('D2', self.color_lab)
        #         worksheet.write('E1', 'Eg')
        #         worksheet.write('F1', 'eff_CM')
        #
        #         worksheet.write('G1', 'Color')
        #         worksheet.write('G2', self.color_2005)
        #         worksheet.write('H1', 'Method')
        #         worksheet.write('H2', 'D2')
        #
        #         worksheet.write('I1', 'Center')
        #         worksheet.write('I2', center)
        #         worksheet.write('J1', 'Width')
        #         worksheet.write('J2', width)
        #
        #         worksheet.write('K1', 'Center2')
        #         worksheet.write('K2', center2)
        #         worksheet.write('L1', 'Width2')
        #         worksheet.write('L2', width2)
        #
        #         workbook.close()
        #return delta_E
        return scaled_E

    
    def op_delta_E_2gauss(self,x, df=df):
        center = x[0]
        width = x[1]
        peak1 = x[2]
        center2 = x[3]
        width2 = x[4]
        peak2 = x[5]
        spectrum = gen_spectrum_2gauss(center,width,peak1,center2,width2,peak2)
        XYZ=spec_to_xyz(spectrum, df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        Lab=convert_color(XYZ, LabColor)

        delta_E=delta_E_CIE2000((Lab.lab_l,Lab.lab_a,Lab.lab_b),self.target_color_LAB)
        if delta_E < self.dE_threshold:
            mean_R = np.mean(spectrum)
            scaled_E = mean_R*delta_E # penalize high average R
        else:
            scaled_E = delta_E
        #print('delta_E 2gauss is ',delta_E)
        # if delta_E < self.dE_threshold:
        #     for i in range(2):
        #         xcel_file = self.color_2005+'_G2_'+'dE_'+str(delta_E).replace('.','_')#color_name
        #         cwd = os.getcwd()
        #         filepath=os.path.join(cwd, r"color_data",xcel_file+'.xlsx')
        #
        #         #print('Delta E Below Threshold')
        #         cp_list = ([center, width, center2, width2,spectrum,Lab,filepath])
        #         self.g2_cp[delta_E] = cp_list
        #
        #
        #         workbook = xlsxwriter.Workbook(filepath)
        #         worksheet = workbook.add_worksheet()
        #         worksheet.write('A1', 'Efficiency')
        #         worksheet.write('B1', 'dE_G2')
        #         worksheet.write('B2', delta_E)
        #         worksheet.write('C1', 'RGB color')
        #         worksheet.write('C2',self.color)
        #         worksheet.write('D1','LAB color')
        #         worksheet.write('D2', self.color_lab)
        #         worksheet.write('E1', 'Eg')
        #         worksheet.write('F1', 'eff_CM')
        #
        #         worksheet.write('G1', 'Color')
        #         worksheet.write('G2', self.color_2005)
        #         worksheet.write('H1', 'Method')
        #         worksheet.write('H2', 'G2')
        #
        #         worksheet.write('I1', 'Center')
        #         worksheet.write('I2', center)
        #         worksheet.write('J1', 'Width')
        #         worksheet.write('J2', width)
        #
        #         worksheet.write('K1', 'Center2')
        #         worksheet.write('K2', center2)
        #         worksheet.write('L1', 'Width2')
        #         worksheet.write('L2', width2)
        #
        #         workbook.close()
        # return delta_E
        return scaled_E


class op_1dip_MJ():

    def __init__(self, tg, th):

        self.target_color_LAB = tg
        self.dE_threshold1d = th

    def evaluate(self,x, df=df):

        center = x[0]
        width = x[1]
        peak = x[2]
        Eg1 = x[3]
        Eg2 = x[4]

        db_junction0 = Junction(kind='DB', T=293, Eg=Eg1, A=1, R_shunt=np.inf, n=3.5)
        db_junction1 = Junction(kind='DB', T=293, Eg=Eg2, A=1, R_shunt=np.inf, n=3.5)

        solar_cell = SolarCell([db_junction1, db_junction0], T=293, R_series=0)

        #base = x[3]
        spectrum = gen_spectrum_1dip(center,width,peak,base=0)
        XYZ=spec_to_xyz(spectrum, df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        Lab=convert_color(XYZ, LabColor)
        delta_E=delta_E_CIE2000((Lab.lab_l,Lab.lab_a,Lab.lab_b),self.target_color_LAB)

        if delta_E < self.dE_threshold1d:

            spec_cell = 1 - gen_spectrum_1dip(center, width, peak, wl_cell)

            light_for_cells = LightSource(source_type='custom', x_data=wl_cell,
                                          y_data=light.spectrum(wl_cell)[1] * spec_cell,
                                          input_units='photon_flux_per_nm', output_units='photon_flux_per_nm')

            # plt.figure()
            # plt.plot(wl, initial_light_source.spectrum(wl)[1] * (1-spectrum))
            # plt.title(str(Eg1) + str(Eg2))
            # plt.show()

            V = np.linspace(0, 5, 500)

            solar_cell_solver(solar_cell, 'iv',
                              user_options={'T_ambient': 293, 'db_mode': 'top_hat',
                                            'voltages': V,
                                            'light_iv': True,
                                            'internal_voltages': np.linspace(-1, 5, 1100),
                                            'wavelength': wl_cell * 1e-9,
                                            'mpp': True,
                                            'light_source': light_for_cells})
            eta = solar_cell.iv['Eta']
            print(eta*100)

            if eta > 0.45:
                print('WEIRD EFFICIENCY', x)

            return -eta

        else:
            return delta_E

    def plot(self, x, df=df):


        center = x[0]
        width = x[1]
        peak = x[2]
        Eg1 = x[3]
        Eg2 = x[4]

        db_junction0 = Junction(kind='DB', T=293, Eg=Eg1, A=1, R_shunt=np.inf, n=3.5)
        db_junction1 = Junction(kind='DB', T=293, Eg=Eg2, A=1, R_shunt=np.inf, n=3.5)

        solar_cell = SolarCell([db_junction1, db_junction0], T=293, R_series=0)

        #base = x[3]
        spectrum = gen_spectrum_1dip(center,width,peak,base=0)
        XYZ=spec_to_xyz(spectrum, df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        Lab=convert_color(XYZ, LabColor)
        delta_E=delta_E_CIE2000((Lab.lab_l,Lab.lab_a,Lab.lab_b),self.target_color_LAB)

        spec_cell = 1 - gen_spectrum_1dip(center, width, peak, wl_cell)

        light_for_cells = LightSource(source_type='custom', x_data=wl_cell,
                                      y_data=light.spectrum(wl_cell)[1] * spec_cell,
                                      input_units='photon_flux_per_nm', output_units='photon_flux_per_nm')

        # plt.figure()
        # plt.plot(wl, initial_light_source.spectrum(wl)[1] * (1-spectrum))
        # plt.title(str(Eg1) + str(Eg2))
        # plt.show()

        V = np.linspace(0, 5, 500)

        solar_cell_solver(solar_cell, 'iv',
                          user_options={'T_ambient': 293, 'db_mode': 'top_hat',
                                        'voltages': V,
                                        'light_iv': True,
                                        'internal_voltages': np.linspace(-1, 5, 1100),
                                        'wavelength': wl_cell * 1e-9,
                                        'mpp': True,
                                        'light_source': light_for_cells})

        eta = solar_cell.iv['Eta']

        plt.figure()
        plt.plot(V, solar_cell.iv.IV[1], 'k', linewidth=4, label='Total')
        plt.plot(V, -solar_cell[0].iv(V), 'r', label='Bottom')
        plt.plot(V, -solar_cell[1].iv(V), 'g', label='Middle')
        plt.ylim(0, 400)
        plt.xlim(0, 3.75)
        plt.legend()
        plt.xlabel('Voltage (V)')
        plt.ylabel('Current (A/m$^2$)')
        plt.text(3, 150, str(np.round(eta*100, 2)))
        plt.show()

        plt.figure()
        plt.plot(wl_cell, light.spectrum(wl_cell)[1] * spec_cell)
        plt.show()



class op_2dip_MJ():

    def __init__(self, tg, th):

        self.target_color_LAB = tg
        self.dE_threshold = th

    def evaluate(self,x, df=df):

        center = x[0]
        width = x[1]
        peak = x[2]
        center2 = x[3]
        width2 = x[4]
        peak2 = x[5]
        Eg1 = x[6]
        Eg2 = x[7]

        db_junction0 = Junction(kind='DB', T=293, Eg=Eg1, A=1, R_shunt=np.inf, n=3.5)
        db_junction1 = Junction(kind='DB', T=293, Eg=Eg2, A=1, R_shunt=np.inf, n=3.5)

        solar_cell = SolarCell([db_junction1, db_junction0], T=293, R_series=0)

        #base = x[3]
        spectrum = gen_spectrum_2dip(center,width,peak,center2, width2, peak2,base=0)
        XYZ=spec_to_xyz(spectrum, df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        Lab=convert_color(XYZ, LabColor)
        delta_E=delta_E_CIE2000((Lab.lab_l,Lab.lab_a,Lab.lab_b),self.target_color_LAB)

        if delta_E < self.dE_threshold:

            spec_cell = 1 - gen_spectrum_2dip(center,width,peak,center2, width2, peak2, wl_cell)

            light_for_cells = LightSource(source_type='custom', x_data=wl_cell,
                                          y_data=light.spectrum(wl_cell)[1] * spec_cell,
                                          input_units='photon_flux_per_nm', output_units='photon_flux_per_nm')

            # plt.figure()
            # plt.plot(wl, initial_light_source.spectrum(wl)[1] * (1-spectrum))
            # plt.title(str(Eg1) + str(Eg2))
            # plt.show()

            V = np.linspace(0, 5, 500)

            solar_cell_solver(solar_cell, 'iv',
                              user_options={'T_ambient': 293, 'db_mode': 'top_hat',
                                            'voltages': V,
                                            'light_iv': True,
                                            'internal_voltages': np.linspace(-1, 5, 1100),
                                            'wavelength': wl_cell * 1e-9,
                                            'mpp': True,
                                            'light_source': light_for_cells})
            eta = solar_cell.iv['Eta']
            print(eta*100)

            if eta > 0.45:
                print('WEIRD EFFICIENCY', x)

            return -eta

        else:
            return delta_E

    def plot(self, x, df=df):

        center = x[0]
        width = x[1]
        peak = x[2]
        center2 = x[3]
        width2 = x[4]
        peak2 = x[5]
        Eg1 = x[6]
        Eg2 = x[7]

        db_junction0 = Junction(kind='DB', T=293, Eg=Eg1, A=1, R_shunt=np.inf, n=3.5)
        db_junction1 = Junction(kind='DB', T=293, Eg=Eg2, A=1, R_shunt=np.inf, n=3.5)

        solar_cell = SolarCell([db_junction1, db_junction0], T=293, R_series=0)

        #base = x[3]
        spectrum = gen_spectrum_2dip(center,width,peak,center2, width2, peak2,base=0)
        XYZ=spec_to_xyz(spectrum, df)
        XYZ=XYZColor(XYZ[0],XYZ[1],XYZ[2])
        Lab=convert_color(XYZ, LabColor)
        delta_E=delta_E_CIE2000((Lab.lab_l,Lab.lab_a,Lab.lab_b),self.target_color_LAB)

        spec_cell = 1 - gen_spectrum_2dip(center,width,peak,center2, width2, peak2, wl_cell)

        light_for_cells = LightSource(source_type='custom', x_data=wl_cell,
                                      y_data=light.spectrum(wl_cell)[1] * spec_cell,
                                      input_units='photon_flux_per_nm', output_units='photon_flux_per_nm')

        # plt.figure()
        # plt.plot(wl, initial_light_source.spectrum(wl)[1] * (1-spectrum))
        # plt.title(str(Eg1) + str(Eg2))
        # plt.show()

        V = np.linspace(0, 5, 500)

        solar_cell_solver(solar_cell, 'iv',
                          user_options={'T_ambient': 293, 'db_mode': 'top_hat',
                                        'voltages': V,
                                        'light_iv': True,
                                        'internal_voltages': np.linspace(-1, 5, 1100),
                                        'wavelength': wl_cell * 1e-9,
                                        'mpp': True,
                                        'light_source': light_for_cells})

        eta = solar_cell.iv['Eta']

        plt.figure()
        plt.plot(V, solar_cell.iv.IV[1], 'k', linewidth=4, label='Total')
        plt.plot(V, -solar_cell[0].iv(V), 'r', label='Bottom')
        plt.plot(V, -solar_cell[1].iv(V), 'g', label='Middle')
        plt.ylim(0, 400)
        plt.xlim(0, 3.75)
        plt.legend()
        plt.xlabel('Voltage (V)')
        plt.ylabel('Current (A/m$^2$)')
        plt.text(3, 150, str(np.round(eta*100, 2)))
        plt.show()

        plt.figure()
        plt.plot(wl_cell, light.spectrum(wl_cell)[1] * spec_cell)
        plt.show()

